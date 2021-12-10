import numpy
import datetime	                #Used numpy to round off the underlying price to the nearest hundred as the strike price is discrete
from nsepy import get_history   #function imported to get index price data and option price data from nse website
from nsepy.derivatives import get_expiry_date
import openpyxl
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait 
from selenium.webdriver.support import expected_conditions as EC

file_location = '/home/arvind/summer_sweta/financial_math/computation/option.xlsx'
wb = openpyxl.load_workbook(file_location)
ws=wb.active
user_name=input("Enter the name you want to show on the excel sheet?")

underlying_price_column=4
strike_column=6
maturity_days_column=9
option_price_column=7
imp_vol_column=10

browser = webdriver.Firefox(executable_path=r'/usr/local/bin/geckodriver')
print(type(browser))
browser.get("https://www.option-price.com/implied-volatility.php")


sxyear = int(input('Enter the starting expiry year(e.g. 2013)'))    #to get the starting option expiry year from the user
sxmonth = int(input('Enter the starting expiry month(e.g. 6,12)'))  #to get the starting option expiry month from the user
							        	
exyear = int(input('Enter the ending expiry year(e.g. 2013)'))      #to get the starting option expiry year from the user
exmonth = int(input('Enter the ending expiry month(e.g. 6,12)'))    #to get the starting option expiry month from the user

row_counter=2                                                                 #'i' is the counter for rows of the excel sheet
while(True):
	expiry=get_expiry_date(year=sxyear,month=sxmonth)           #to get the expiry date for the given year and month
	sdate =expiry-(datetime.timedelta(days=60)) 
	edate =expiry-(datetime.timedelta(days=40))		    #to keep the maturity days between 40 to 60 days

	while(sdate!=edate):	  							
		nifty_price = get_history(symbol="NIFTY 50",        #to get the index price on the particular date
		start=sdate,
		end=sdate,
		index=True)

	
		if (nifty_price.empty):				          #exception handling
			sdate += datetime.timedelta(days=1)
			continue
		
		nifty_option = get_history(symbol="NIFTY",            #to get index option price
		start=sdate,
		end=sdate,
		index=True,
		option_type='CE',
		strike_price=int(numpy.round(nifty_price.get('Close'),-2)),
		expiry_date=expiry)		              	
		
		
		if (nifty_option.empty):			            #exception handling
			sdate += datetime.timedelta(days=1)
			continue
		
		
		if (int(nifty_option.get('Number of Contracts'))):	    #we want the data only of days when the option was traded
#			user_name
			symbol=nifty_option.get('Symbol').tolist()[0]  #Converting the element to list and then strings
#		 	sdate
			underlying_price=nifty_price.get('Close').tolist()[0]
#			expiry
			strike=nifty_option.get('Strike Price').tolist()[0]
			option_price=nifty_option.get('Settle Price').tolist()[0]
			contracts_no=nifty_option.get('Number of Contracts').tolist()[0]
			maturity_days=(expiry-sdate).days
			data=[user_name,symbol,sdate,underlying_price,expiry,strike,option_price,contracts_no,maturity_days]
			
			column_counter=1                                        #'j' keeps track of the column
			for c in data:
				ws.cell(row=row_counter,column=column_counter).value=c
				column_counter +=1

			


			up = browser.find_element_by_id('up')
			up.clear()   #Clears the existing field
			up.send_keys(str(underlying_price))   #Sends the value

			ex = browser.find_element_by_id('ex')
			ex.clear()
			ex.send_keys(str(strike))

			time = browser.find_element_by_id('time')
			time.clear()
			time.send_keys(str(maturity_days))

			rate = browser.find_element_by_id('rate')
			rate.clear()
			rate.send_keys('0')

			div = browser.find_element_by_id('div')
			div.clear()
			div.send_keys('0')

			mp = browser.find_element_by_id('marketprice')
			mp.clear()
			mp.send_keys(str(option_price))


			sub = browser.find_element_by_name('Calculate')
			sub.click()
    	
			imp_vol = browser.find_element_by_id('result').get_attribute("value")
			ws.cell(row = row_counter, column = column_counter).value = imp_vol   #Passing the results in a sheet
			

			row_counter +=1
			sdate += datetime.timedelta(days=1)
	if(sxyear==exyear and sxmonth==exmonth):                  #this condtion is to come out of the loop after the starting expiry 	     		 							year and month are equal with the ending expiry year and month
		break;
	if(sxmonth !=12):                                        #if...else is to increment the sxmonth and sxyear
		sxmonth +=1
	else:
		sxyear +=1
		sxmonth=1
    
wb.save(file_location)						#save the file once it is done














